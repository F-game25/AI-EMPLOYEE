"""document_extractor — extract text from any supported file type.

Usage::

    from tools.document_extractor import extract

    result = extract("/path/to/file.pdf")
    # Returns: { text, metadata, pages, language, file_type }
    # Images:  { text: "", file_type: "image", needs_vision: True, metadata: {width, height} }
    # Errors:  { error: "..." }
"""
from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import docx
except ImportError:
    docx = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from PIL import Image as PilImage
except ImportError:
    PilImage = None

# Map extension → language label for code files
_CODE_LANGUAGES = {
    ".py": "python", ".js": "javascript", ".ts": "typescript",
    ".jsx": "javascript", ".tsx": "typescript", ".rs": "rust",
    ".toml": "toml", ".sh": "bash", ".css": "css", ".html": "html",
    ".json": "json", ".md": "markdown",
}

_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}


def _strip_html(text: str) -> str:
    """Remove HTML tags and collapse whitespace."""
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s{2,}", " ", text).strip()


def _extract_plain(path: Path, strip_html: bool = False) -> dict[str, Any]:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        if strip_html:
            text = _strip_html(text)
        return {"text": text, "pages": 1, "language": "text", "file_type": path.suffix.lstrip(".")}
    except Exception as e:
        return {"error": f"Failed to read file: {e}"}


def _extract_csv(path: Path) -> dict[str, Any]:
    try:
        lines = []
        with path.open(encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            if headers:
                lines.append("Headers: " + ", ".join(headers))
            row_count = 0
            for row in reader:
                if row_count >= 500:
                    break
                lines.append(", ".join(row))
                row_count += 1
        return {
            "text": "\n".join(lines),
            "pages": 1,
            "language": "csv",
            "file_type": "csv",
            "metadata": {"headers": headers, "rows_sampled": row_count},
        }
    except Exception as e:
        return {"error": f"Failed to parse CSV: {e}"}


def _extract_pdf(path: Path) -> dict[str, Any]:
    if pdfplumber is None:
        return {"error": "pdfplumber not installed — run: pip install pdfplumber"}
    try:
        pages_text = []
        with pdfplumber.open(str(path)) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                t = page.extract_text() or ""
                pages_text.append(f"\n---page {i}---\n{t}")
        return {
            "text": "\n".join(pages_text),
            "pages": len(pages_text),
            "language": "text",
            "file_type": "pdf",
        }
    except Exception as e:
        return {"error": f"Failed to parse PDF: {e}"}


def _extract_docx(path: Path) -> dict[str, Any]:
    if docx is None:
        return {"error": "python-docx not installed — run: pip install python-docx"}
    try:
        doc = docx.Document(str(path))
        lines = []
        for para in doc.paragraphs:
            style = para.style.name if para.style else ""
            prefix = "# " if style.startswith("Heading") else ""
            if para.text.strip():
                lines.append(prefix + para.text.strip())
        return {
            "text": "\n".join(lines),
            "pages": 1,
            "language": "text",
            "file_type": "docx",
        }
    except Exception as e:
        return {"error": f"Failed to parse DOCX: {e}"}


def _extract_xlsx(path: Path) -> dict[str, Any]:
    if openpyxl is None:
        return {"error": "openpyxl not installed — run: pip install openpyxl"}
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"=== Sheet: {sheet_name} ===")
            rows_written = 0
            for row in ws.iter_rows(values_only=True):
                if rows_written >= 200:
                    break
                parts.append(", ".join("" if v is None else str(v) for v in row))
                rows_written += 1
        wb.close()
        return {
            "text": "\n".join(parts),
            "pages": len(wb.sheetnames),
            "language": "spreadsheet",
            "file_type": "xlsx",
            "metadata": {"sheets": wb.sheetnames},
        }
    except Exception as e:
        return {"error": f"Failed to parse XLSX: {e}"}


def _extract_image(path: Path) -> dict[str, Any]:
    meta: dict[str, Any] = {}
    if PilImage is not None:
        try:
            with PilImage.open(str(path)) as img:
                meta["width"], meta["height"] = img.size
                meta["mode"] = img.mode
        except Exception:
            pass

    try:
        from tools.vision_analyzer import analyze_image  # noqa: PLC0415
        result = analyze_image(str(path))
        result.setdefault("pages", 1)
        result.setdefault("language", "image")
        result.setdefault("metadata", {}).update(meta)
        return result
    except Exception as e:
        logger.warning("vision_analyzer unavailable: %s", e)
        return {
            "text": "",
            "file_type": "image",
            "needs_vision": True,
            "pages": 0,
            "language": "image",
            "metadata": meta,
        }


def extract(file_path: str) -> dict[str, Any]:
    """Extract text from a file. Returns dict with text, metadata, pages, language, file_type."""
    path = Path(file_path)
    ext = path.suffix.lower()

    if not path.exists():
        return {"error": f"File not found: {file_path}"}

    # Images
    if ext in _IMAGE_EXTS:
        return _extract_image(path)

    # PDF
    if ext == ".pdf":
        result = _extract_pdf(path)
    elif ext == ".docx":
        result = _extract_docx(path)
    elif ext == ".xlsx":
        result = _extract_xlsx(path)
    elif ext == ".csv":
        result = _extract_csv(path)
    elif ext == ".html":
        result = _extract_plain(path, strip_html=True)
    elif ext in _CODE_LANGUAGES:
        result = _extract_plain(path)
        if "error" not in result:
            result["language"] = _CODE_LANGUAGES[ext]
    else:
        # Unknown: attempt plain text, fallback to empty
        try:
            result = _extract_plain(path)
        except Exception:
            result = {"text": "", "pages": 0, "language": "unknown", "file_type": ext.lstrip(".")}

    # Attach metadata if missing
    if "error" not in result:
        result.setdefault("metadata", {})
        result["metadata"].setdefault("file_name", path.name)
        result["metadata"].setdefault("file_size", path.stat().st_size)
        result.setdefault("pages", 1)
        result.setdefault("language", "text")
        result.setdefault("file_type", ext.lstrip("."))

    return result
